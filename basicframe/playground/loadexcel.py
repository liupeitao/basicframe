import string

from openpyxl import load_workbook

# 加载Excel文件
# wb = load_workbook(filename='/home/ptking/多语种-文本-20230615.xlsx')
#
# # 获取第一个工作表
# sheet = wb.active
#
# # 获取列名映射关系
# column_names = {}
# for col_index, col_letter in enumerate(string.ascii_uppercase, start=1):
#     column_names[col_index] = col_letter
# cursor_dict = {}
# # 遍历每一行
# for row in sheet.iter_rows():
#     row_data = []
#     # 遍历每一列
#     for cell in row:
#         # 获取单元格的行和列索引
#         row_index = cell.row
#         col_index = cell.column
#         # 获取单元格的值
#         cell_value = cell.value
#         # 判断是否为合并单元格
#         is_merged_cell = False
#         merged_value = None
#         for merged_cell_range in sheet.merged_cells.ranges:
#             min_row, min_col, max_row, max_col = merged_cell_range.min_row, merged_cell_range.min_col, merged_cell_range.max_row, merged_cell_range.max_col
#             if min_row <= row_index <= max_row and min_col <= col_index <= max_col:
#                 is_merged_cell = True
#                 merged_value = sheet.cell(row=min_row, column=min_col).value
#                 break
#         # 输出单元格信息
#         if is_merged_cell and merged_value is not None:
#             cell_value = merged_value
#         # 构建包含列名的字典
#         col_name = column_names.get(col_index, str(col_index))
#         cell_data = {"列": col_name, "value": cell_value}
#         res_dcit.update(cell_data)
#     yield cursor_dict
    # new_dict = {}
    # for item in row_data:
    #     column = item['列']
    #     value = item['value']
    #     new_dict[column] = value
    # print(new_dict)

def get_one_line(file_name):
    wb = load_workbook(file_name)
    # 获取第一个工作表
    sheet = wb.active

    # 获取列名映射关系
    column_names = {}
    for col_index, col_letter in enumerate(string.ascii_uppercase, start=1):
        column_names[col_index] = col_letter
    cursor_dict = {}
    # 遍历每一行
    for row in sheet.iter_rows():
        row_data = {}
        # 遍历每一列
        for cell in row:
            # 获取单元格的行和列索引
            row_index = cell.row
            col_index = cell.column
            # 获取单元格的值
            cell_value = cell.value
            # 判断是否为合并单元格
            is_merged_cell = False
            merged_value = None
            for merged_cell_range in sheet.merged_cells.ranges:
                min_row, min_col, max_row, max_col = merged_cell_range.min_row, merged_cell_range.min_col, merged_cell_range.max_row, merged_cell_range.max_col
                if min_row <= row_index <= max_row and min_col <= col_index <= max_col:
                    is_merged_cell = True
                    merged_value = sheet.cell(row=min_row, column=min_col).value
                    break
            # 输出单元格信息
            if is_merged_cell and merged_value is not None:
                cell_value = merged_value
            # 构建包含列名的字典
            col_name = column_names.get(col_index, str(col_index))
            cell_data = {"列": col_name, "value": cell_value}
            row_data.append(cell_data)
        yield row_data


file_name = '/home/ptking/多语种-文本-20230615.xlsx'
for i in get_one_line(file_name):
    print(i)